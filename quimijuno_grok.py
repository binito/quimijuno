import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys
import logging

# Configuração básica de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class PDFToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor de Cotação PDF para Excel")
        self.center_window(440, 260)  # Centraliza a janela

        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()

        ttk.Label(main_frame, text="Ficheiro PDF de Cotação:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.pdf_path, width=50).grid(row=1, column=0, padx=5)
        self.btn_browse_pdf = ttk.Button(main_frame, text="Procurar", command=self.browse_pdf)
        self.btn_browse_pdf.grid(row=1, column=1)

        ttk.Label(main_frame, text="Destino do ficheiro Excel:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.excel_path, width=50).grid(row=3, column=0, padx=5)
        self.btn_browse_excel = ttk.Button(main_frame, text="Procurar", command=self.browse_excel)
        self.btn_browse_excel.grid(row=3, column=1)

        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.grid(row=4, column=0, columnspan=2, pady=20)

        self.btn_convert = ttk.Button(main_frame, text="Converter", command=self.convert)
        self.btn_convert.grid(row=5, column=0, columnspan=2, pady=10)

        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=6, column=0, columnspan=2)

    def center_window(self, width, height):
        """Centraliza a janela na tela."""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def toggle_buttons(self, state):
        """Habilita ou desabilita os botões da interface."""
        self.btn_browse_pdf.config(state=state)
        self.btn_browse_excel.config(state=state)
        self.btn_convert.config(state=state)

    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Selecione o ficheiro PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        if filename:
            self.pdf_path.set(filename)
            excel_name = os.path.splitext(os.path.basename(filename))[0] + ".xlsx"
            self.excel_path.set(os.path.join(os.path.dirname(filename), excel_name))

    def browse_excel(self):
        filename = filedialog.asksaveasfilename(
            title="Salvar ficheiro Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.excel_path.set(filename)

    def extrair_texto_pdf(self, caminho_pdf):
        """Extrai texto de um arquivo PDF."""
        try:
            with open(caminho_pdf, "rb") as f:
                leitor = PdfReader(f)
                texto = ""
                for pagina in leitor.pages:
                    texto_pagina = pagina.extract_text()
                    if texto_pagina:
                        texto += texto_pagina + "\n"
                if not texto.strip():
                    raise ValueError("Nenhum texto extraível encontrado no PDF.")
                return texto
        except Exception as e:
            logging.error(f"Erro ao ler o PDF: {e}")
            raise Exception(f"Não foi possível ler o arquivo PDF: {e}")

    def extrair_dados(self, texto):
        """Extrai dados do texto usando expressões regulares."""
        produtos = []
        if "DESCRIÇÃO" not in texto:
            logging.warning("Cabeçalho 'DESCRIÇÃO' não encontrado.")
            messagebox.showwarning("Aviso", "Cabeçalho 'DESCRIÇÃO' não encontrado.")
            return produtos

        texto = texto.split("DESCRIÇÃO", 1)[1]
        linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
        texto = "\n".join(linhas)

        padrao = re.compile(
            r"(?P<referencia>\[[^\]]+\])\s*"
            r"(?P<descricao>(?:(?!\[\w+\]).)*?)\s*"
            r"(?P<quantidade>\d+[.,]\d+)"
            r"(?:\s*(?P<unidade>[kK][gG]|[lL](?:itros?)?|UN))?"
            r"\s+(?P<preco>\d+[.,]\d+)"
            r"\s+(?P<impostos>IVA\s*\d+%?)"
            r"\s+(?P<amount>\d+[.,]\d+\s*€)",
            flags=re.DOTALL | re.IGNORECASE
        )

        for match in padrao.finditer(texto):
            referencia = match.group("referencia").strip('[]')
            descricao = " ".join(match.group("descricao").split())
            quantidade = float(match.group("quantidade").replace(',', '.'))
            unidade = (match.group("unidade") or "").upper()
            unidade = {'KG': 'KG', 'LITROS': 'L', 'L': 'L', 'UN': 'Unidades'}.get(unidade[:2], unidade)
            preco = float(match.group("preco").replace(',', '.'))
            impostos_text = match.group("impostos")
            impostos = float(re.search(r'\d+[.,]?\d*', impostos_text).group().replace(',', '.')) / 100.0 if re.search(r'\d+', impostos_text) else 0.0
            amount = float(match.group("amount").replace('€', '').replace(',', '.').strip())

            produtos.append({
                "REFERÊNCIA": referencia,
                "DESCRIÇÃO": descricao,
                "QUANTIDADE": quantidade,
                "UNIDADE": unidade,
                "PREÇO UNITÁRIO": preco,
                "IMPOSTOS": impostos,
                "AMOUNT": amount
            })

        if not produtos:
            logging.warning("Nenhum produto encontrado no texto.")
            messagebox.showwarning("Aviso", "Nenhum produto encontrado com o padrão definido.")
        return produtos

    def escrever_excel(self, produtos, caminho_excel):
        """Escreve os dados extraídos em um arquivo Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotação"
        wb.properties.creator = "Comsoftweb"

        cabecalhos = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
        header_style = {'font': Font(bold=True, color="FFFFFF"), 'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"), 'alignment': Alignment(horizontal="center")}

        # Escreve cabeçalhos
        ws.append(cabecalhos)
        for col, cell in enumerate(ws[1], 1):
            for attr, value in header_style.items():
                setattr(cell, attr, value)

        # Escreve dados
        for produto in produtos:
            ws.append([produto[header] for header in cabecalhos])

        # Ajusta largura das colunas
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = max_length

        wb.save(caminho_excel)
        logging.info(f"Arquivo Excel salvo em: {caminho_excel}")

    def convert(self):
        """Converte o PDF para Excel."""
        if not self.pdf_path.get() or not self.excel_path.get():
            messagebox.showerror("Erro", "Por favor, selecione os ficheiros de entrada e saída.")
            return

        self.toggle_buttons('disabled')
        try:
            self.status_label.config(text="A processar PDF...")
            self.progress['value'] = 33
            self.root.update_idletasks()

            texto = self.extrair_texto_pdf(self.pdf_path.get())

            self.status_label.config(text="A extrair dados...")
            self.progress['value'] = 66
            self.root.update_idletasks()

            produtos = self.extrair_dados(texto)

            if produtos:
                self.status_label.config(text="A criar ficheiro Excel...")
                self.progress['value'] = 90
                self.root.update_idletasks()

                self.escrever_excel(produtos, self.excel_path.get())

                self.progress['value'] = 100
                self.status_label.config(text="Conversão concluída com sucesso!")
                messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {self.excel_path.get()}")
            else:
                self.status_label.config(text="Nenhum produto encontrado.")
                self.progress['value'] = 0

        except Exception as e:
            logging.error(f"Erro durante a conversão: {e}")
            self.status_label.config(text="Erro durante a conversão")
            self.progress['value'] = 0
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")
        finally:
            self.toggle_buttons('normal')

def main():
    root = tk.Tk()
    icone_path = os.path.join(sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.abspath(os.path.dirname(__file__)), 'comsoftweb.ico')
    try:
        root.iconbitmap(icone_path)
    except Exception as e:
        logging.warning(f"Não foi possível carregar o ícone: {e}")
    
    app = PDFToExcelConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()