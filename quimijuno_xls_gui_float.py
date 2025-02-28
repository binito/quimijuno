import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

class PDFToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor de Cotação PDF para Excel")
        self.root.geometry("600x300")
        
        # Criar frame principal
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Variáveis para armazenar caminhos
        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        
        # Componentes da GUI
        # PDF Input
        ttk.Label(main_frame, text="Ficheiro PDF de Cotação:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.pdf_path, width=50).grid(row=1, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_pdf).grid(row=1, column=1)
        
        # Excel Output
        ttk.Label(main_frame, text="Destino do ficheiro Excel:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.excel_path, width=50).grid(row=3, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_excel).grid(row=3, column=1)
        
        # Progress Bar
        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.grid(row=4, column=0, columnspan=2, pady=20)
        
        # Botão Converter
        ttk.Button(main_frame, text="Converter", command=self.convert).grid(row=5, column=0, columnspan=2, pady=10)
        
        # Label de status
        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=6, column=0, columnspan=2)

    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Selecione o ficheiro PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        if filename:
            self.pdf_path.set(filename)
            # Define automaticamente o caminho do Excel
            excel_name = os.path.splitext(os.path.basename(filename))[0] + ".xlsx"
            self.excel_path.set(os.path.join(os.path.dirname(filename), excel_name))

    def browse_excel(self):
        filename = filedialog.asksaveasfilename(
            title="Salvar ficheiro Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.excel_path.set(filename)

    def extrair_texto_pdf(self, caminho_pdf):
        texto = ""
        # Abre o ficheiro PDF em modo binário e extrai o texto de cada página
        with open(caminho_pdf, "rb") as f:
            leitor = PdfReader(f)
            for pagina in leitor.pages:
                texto += pagina.extract_text() + "\n"
        return texto

    def extrair_dados(self, texto):
        produtos = []
        
        # Separa o texto a partir do cabeçalho "DESCRIÇÃO"
        if "DESCRIÇÃO" in texto:
            texto = texto.split("DESCRIÇÃO", 1)[1]
        else:
            messagebox.showwarning("Aviso", "Cabeçalho 'DESCRIÇÃO' não encontrado.")
            return []
        
        # Remove linhas vazias e espaços desnecessários
        linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
        texto = "\n".join(linhas)
        
        # Expressão regular atualizada para:
        # - Capturar referência, descrição, quantidade, unidade (KG, L, UN), preço, impostos e montante
        # - A unidade é opcional; se não existir, ficará em branco.
        padrao = re.compile(
            r"(?P<referencia>\[[^\]]+\])\s*"  # Captura a referência entre colchetes
            r"(?P<descricao>(?:(?!\[\w+\]).)*?)(?=\s*\d+[.,]\d+(?:\s*(?:[kK][gG]|[lL](?:itros?)?|UN))?)\s*"  # Captura a descrição
            r"(?P<quantidade>\d+[.,]\d+)"     # Captura a quantidade
            r"(?:\s*(?P<unidade>(?:[kK][gG]|[lL](?:itros?)?|UN)))?"  # Captura opcionalmente a unidade
            r"\s+(?P<preco>\d+[.,]\d+)"       # Captura o preço unitário
            r"\s+(?P<impostos>IVA\s*\d+%?)"    # Captura os impostos (ex: IVA 23% ou IVA 23)
            r"\s+(?P<amount>\d+[.,]\d+\s*€)", # Captura o montante total
            flags=re.DOTALL | re.IGNORECASE
        )
        
        matches = list(padrao.finditer(texto))
        
        if not matches:
            messagebox.showwarning("Aviso", "Nenhuma entrada de produto foi encontrada com o padrão definido.")
        
        for match in matches:
            # Processa a referência removendo os colchetes
            referencia = match.group("referencia").strip('[]')
            # Remove espaços extra na descrição
            descricao = " ".join(match.group("descricao").split())
            
            # Converter a quantidade para float (substituindo a vírgula por ponto)
            quantidade_str = match.group("quantidade").strip()
            quantidade = float(quantidade_str.replace(',', '.'))
            
            # Processa a unidade (opcional)
            unidade = match.group("unidade")
            if unidade is None:
                unidade = ""  # Se não detetada, deixa em branco
            else:
                unidade = unidade.upper()
                if unidade.startswith('L'):
                    unidade = 'L'
                elif unidade.startswith('K'):
                    unidade = 'KG'
                elif unidade == 'UN':
                    unidade = 'Unidades'
            
            # Converter o preço unitário para float
            preco_str = match.group("preco").strip()
            preco = float(preco_str.replace(',', '.'))
            
            # Processa os impostos:
            # Extrai o valor numérico (ex: "23" ou "23,5"), converte para float e divide por 100 (ex: 23 -> 0.23)
            impostos_text = match.group("impostos")
            num_impostos = re.search(r'\d+[.,]?\d*', impostos_text)
            if num_impostos:
                impostos = float(num_impostos.group().replace(',', '.')) / 100.0
            else:
                impostos = 0.0
            
            # Converter o montante total para float (remove o símbolo do euro)
            amount_str = match.group("amount").replace('€', '').strip()
            amount = float(amount_str.replace(',', '.'))
            
            produtos.append({
                "REFERÊNCIA": referencia,
                "DESCRIÇÃO": descricao,
                "QUANTIDADE": quantidade,      # Valor numérico
                "UNIDADE": unidade,
                "PREÇO UNITÁRIO": preco,         # Valor numérico
                "IMPOSTOS": impostos,            # Valor numérico (ex: 0.23)
                "AMOUNT": amount                 # Valor numérico
            })
        
        return produtos

    def escrever_excel(self, produtos, caminho_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotação"
        
        cabecalhos = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
        
        # Estilo para os cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Escreve os cabeçalhos na primeira linha
        for col, header in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Escreve os dados dos produtos, mantendo a formatação numérica
        for row, produto in enumerate(produtos, 2):
            for col, header in enumerate(cabecalhos, 1):
                cell = ws.cell(row=row, column=col, value=produto[header])
                cell.alignment = Alignment(horizontal="center")
        
        # Ajusta a largura das colunas com base no comprimento do conteúdo
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(caminho_excel)

    def convert(self):
        if not self.pdf_path.get() or not self.excel_path.get():
            messagebox.showerror("Erro", "Por favor, selecione os ficheiros de entrada e saída.")
            return
        
        try:
            self.status_label.config(text="A processar PDF...")
            self.progress['value'] = 33
            self.root.update_idletasks()
            
            texto = self.extrair_texto_pdf(self.pdf_path.get())
            
            self.status_label.config(text="A extrair dados...")
            self.progress['value'] = 66
            self.root.update_idletasks()
            
            produtos = self.extrair_dados(texto)
            
            if produtos:
                self.status_label.config(text="A criar ficheiro Excel...")
                self.progress['value'] = 90
                self.root.update_idletasks()
                
                self.escrever_excel(produtos, self.excel_path.get())
                
                self.progress['value'] = 100
                self.status_label.config(text="Conversão concluída com sucesso!")
                messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {self.excel_path.get()}")
            else:
                self.status_label.config(text="Nenhum produto encontrado.")
                self.progress['value'] = 0
                
        except Exception as e:
            self.status_label.config(text="Erro durante a conversão")
            self.progress['value'] = 0
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")

def main():
    root = tk.Tk()
    app = PDFToExcelConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
