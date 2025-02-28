import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys  # Import necessário para detectar se estamos a correr no exe do PyInstaller

class PDFToExcelConverter:
    def __init__(self, root):
        """
        Classe principal do Conversor de Cotação PDF para Excel.
        Recebe a janela (root) como argumento e constrói toda a interface gráfica.
        """
        self.root = root
        self.root.title("Conversor de Cotação PDF para Excel")
        
        # Dimensões da janela
        self.root.geometry("450x240")
        
        # Frame principal para conter os widgets
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Variáveis para armazenar os caminhos do PDF e do Excel
        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        
        # Label e entrada para o PDF
        ttk.Label(main_frame, text="Ficheiro PDF de Cotação:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.pdf_path, width=50).grid(row=1, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_pdf).grid(row=1, column=1)
        
        # Label e entrada para o Excel
        ttk.Label(main_frame, text="Destino do ficheiro Excel:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.excel_path, width=50).grid(row=3, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_excel).grid(row=3, column=1)
        
        # Barra de progresso
        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.grid(row=4, column=0, columnspan=2, pady=20)
        
        # Botão para iniciar a conversão
        ttk.Button(main_frame, text="Converter", command=self.convert).grid(row=5, column=0, columnspan=2, pady=10)
        
        # Label para mostrar mensagens de estado
        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=6, column=0, columnspan=2)

    def browse_pdf(self):
        """
        Abre um diálogo para selecionar o ficheiro PDF de entrada.
        Ao selecionar, preenche o campo do PDF e sugere um nome para o Excel.
        """
        filename = filedialog.askopenfilename(
            title="Selecione o ficheiro PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        if filename:
            self.pdf_path.set(filename)
            excel_name = os.path.splitext(os.path.basename(filename))[0] + ".xlsx"
            self.excel_path.set(os.path.join(os.path.dirname(filename), excel_name))

    def browse_excel(self):
        """
        Abre um diálogo para definir o caminho onde será guardado o ficheiro Excel.
        """
        filename = filedialog.asksaveasfilename(
            title="Salvar ficheiro Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.excel_path.set(filename)

    def extrair_texto_pdf(self, caminho_pdf):
        """
        Extrai todo o texto de um ficheiro PDF, página a página.
        Retorna uma string com o texto completo.
        """
        texto = ""
        with open(caminho_pdf, "rb") as f:
            leitor = PdfReader(f)
            for pagina in leitor.pages:
                texto += pagina.extract_text() + "\n"
        return texto

    def extrair_dados(self, texto):
        """
        Extrai os dados de produtos (REFERÊNCIA, DESCRIÇÃO, QUANTIDADE, UNIDADE, PREÇO UNITÁRIO, IMPOSTOS, AMOUNT)
        usando expressões regulares. Retorna uma lista de dicionários.
        """
        produtos = []
        
        # Verifica se existe o cabeçalho "DESCRIÇÃO" no texto
        if "DESCRIÇÃO" in texto:
            texto = texto.split("DESCRIÇÃO", 1)[1]
        else:
            messagebox.showwarning("Aviso", "Cabeçalho 'DESCRIÇÃO' não encontrado.")
            return []
        
        # Limpa linhas vazias e junta tudo
        linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
        texto = "\n".join(linhas)
        
        # Expressão regular para capturar os campos
        padrao = re.compile(
            r"(?P<referencia>\[[^\]]+\])\s*"
            r"(?P<descricao>(?:(?!\[\w+\]).)*?)(?=\s*\d+[.,]\d+(?:\s*(?:[kK][gG]|[lL](?:itros?)?|UN))?)\s*"
            r"(?P<quantidade>\d+[.,]\d+)"
            r"(?:\s*(?P<unidade>(?:[kK][gG]|[lL](?:itros?)?|UN)))?"
            r"\s+(?P<preco>\d+[.,]\d+)"
            r"\s+(?P<impostos>IVA\s*\d+%?)"
            r"\s+(?P<amount>\d+[.,]\d+\s*€)",
            flags=re.DOTALL | re.IGNORECASE
        )
        
        matches = list(padrao.finditer(texto))
        
        if not matches:
            messagebox.showwarning("Aviso", "Nenhuma entrada de produto foi encontrada com o padrão definido.")
        
        for match in matches:
            referencia = match.group("referencia").strip('[]')
            descricao = " ".join(match.group("descricao").split())
            
            quantidade_str = match.group("quantidade").strip()
            quantidade = float(quantidade_str.replace(',', '.'))
            
            unidade = match.group("unidade")
            if unidade is None:
                unidade = ""
            else:
                unidade = unidade.upper()
                if unidade.startswith('L'):
                    unidade = 'L'
                elif unidade.startswith('K'):
                    unidade = 'KG'
                elif unidade == 'UN':
                    unidade = 'Unidades'
            
            preco_str = match.group("preco").strip()
            preco = float(preco_str.replace(',', '.'))
            
            impostos_text = match.group("impostos")
            num_impostos = re.search(r'\d+[.,]?\d*', impostos_text)
            if num_impostos:
                impostos = float(num_impostos.group().replace(',', '.')) / 100.0
            else:
                impostos = 0.0
            
            amount_str = match.group("amount").replace('€', '').strip()
            amount = float(amount_str.replace(',', '.'))
            
            produtos.append({
                "REFERÊNCIA": referencia,
                "DESCRIÇÃO": descricao,
                "QUANTIDADE": quantidade,
                "UNIDADE": unidade,
                "PREÇO UNITÁRIO": preco,
                "IMPOSTOS": impostos,
                "AMOUNT": amount
            })
        
        return produtos

    def escrever_excel(self, produtos, caminho_excel):
        """
        Cria e guarda um ficheiro Excel (caminho_excel) com os dados dos produtos.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotação"
        
        wb.properties.creator = "Comsoftweb"

        
        # Cabeçalhos para a folha Excel
        cabecalhos = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
        
        # Estilos para o cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Escrever os cabeçalhos
        for col, header in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Escrever os produtos
        for row, produto in enumerate(produtos, 2):
            for col, header in enumerate(cabecalhos, 1):
                cell = ws.cell(row=row, column=col, value=produto[header])
                cell.alignment = Alignment(horizontal="center")
        
        # Ajustar a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar o ficheiro Excel
        wb.save(caminho_excel)

    def convert(self):
        """
        Função chamada ao clicar no botão 'Converter'. Faz a leitura do PDF,
        extrai os dados e escreve no Excel, atualizando a barra de progresso.
        """
        if not self.pdf_path.get() or not self.excel_path.get():
            messagebox.showerror("Erro", "Por favor, selecione os ficheiros de entrada e saída.")
            return
        
        try:
            self.status_label.config(text="A processar PDF...")
            self.progress['value'] = 33
            self.root.update_idletasks()
            
            # Extrair texto do PDF
            texto = self.extrair_texto_pdf(self.pdf_path.get())
            
            self.status_label.config(text="A extrair dados...")
            self.progress['value'] = 66
            self.root.update_idletasks()
            
            # Extrair dados dos produtos
            produtos = self.extrair_dados(texto)
            
            if produtos:
                self.status_label.config(text="A criar ficheiro Excel...")
                self.progress['value'] = 90
                self.root.update_idletasks()
                
                # Escrever dados no Excel
                self.escrever_excel(produtos, self.excel_path.get())
                
                self.progress['value'] = 100
                self.status_label.config(text="Conversão concluída com sucesso!")
                messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {self.excel_path.get()}")
            else:
                self.status_label.config(text="Nenhum produto encontrado.")
                self.progress['value'] = 0
                
        except Exception as e:
            self.status_label.config(text="Erro durante a conversão")
            self.progress['value'] = 0
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")

def main():
    """
    Função principal que inicia a aplicação.
    É aqui que garantimos que o ícone seja carregado corretamente,
    tanto em modo de desenvolvimento como quando compilado com PyInstaller.
    """
    root = tk.Tk()
    
    # Verifica se está a correr no executável PyInstaller (que gera a pasta _MEIPASS)
    if hasattr(sys, '_MEIPASS'):
        icone_path = os.path.join(sys._MEIPASS, 'comsoftweb.ico')
    else:
        # Caminho absoluto do .ico na pasta do script
        icone_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'comsoftweb.ico')
    
    # Tentar carregar o ícone
    try:
        root.iconbitmap(icone_path)
    except Exception as e:
        print("Aviso: Não foi possível carregar o ícone. Erro:", e)
    
    # Instanciar a aplicação
    app = PDFToExcelConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
