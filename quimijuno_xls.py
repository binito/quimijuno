import re
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extrair_texto_pdf(caminho_pdf):
    texto = ""
    with open(caminho_pdf, "rb") as f:
        leitor = PdfReader(f)
        for pagina in leitor.pages:
            texto += pagina.extract_text() + "\n"
    return texto

def extrair_dados(texto):
    produtos = []
    
    if "DESCRIÇÃO" in texto:
        texto = texto.split("DESCRIÇÃO", 1)[1]
    else:
        print("Cabeçalho 'DESCRIÇÃO' não encontrado.")
        return []
    
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    texto = "\n".join(linhas)
    
    padrao = re.compile(
        r"(?P<referencia>\[[^\]]+\])\s*"
        r"(?P<descricao>(?:(?!\[\w+\]).)*?)(?=\s*\d+[.,]\d+\s*(?:[kK][gG]|[lL]|[lL]itros?))\s*"
        r"(?P<quantidade>\d+[.,]\d+)\s*"
        r"(?P<unidade>(?:[kK][gG]|[lL]|[lL]itros?))\s+"
        r"(?P<preco>\d+[.,]\d+)\s+"
        r"(?P<impostos>IVA\s*\d+%?)\s+"
        r"(?P<amount>\d+[.,]\d+\s*€)",
        flags=re.DOTALL | re.IGNORECASE
    )
    
    matches = list(padrao.finditer(texto))
    
    if not matches:
        print("Nenhuma entrada de produto foi encontrada com o padrão definido.")
    
    for match in matches:
        referencia = match.group("referencia").strip('[]')
        descricao = " ".join(match.group("descricao").split())
        
        quantidade = match.group("quantidade").strip()
        unidade = match.group("unidade").upper()
        if unidade.startswith('L'):
            unidade = 'L'
        elif unidade.upper().startswith('K'):
            unidade = 'KG'
            
        preco = " ".join(match.group("preco").split())
        
        impostos = match.group("impostos")
        percentagem = re.search(r'\d+%', impostos)
        if percentagem:
            impostos = percentagem.group()
            
        amount = " ".join(match.group("amount").split())
        
        produtos.append({
            "REFERÊNCIA": referencia,
            "DESCRIÇÃO": descricao,
            "QUANTIDADE": quantidade,
            "UNIDADE": unidade,
            "PREÇO UNITÁRIO": preco,
            "IMPOSTOS": impostos,
            "AMOUNT": amount
        })
    
    return produtos

def escrever_excel(produtos, caminho_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotação"
    
    # Define cabeçalhos
    cabecalhos = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
    
    # Estilo para cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Escreve cabeçalhos
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Escreve dados
    for row, produto in enumerate(produtos, 2):
        for col, header in enumerate(cabecalhos, 1):
            cell = ws.cell(row=row, column=col, value=produto[header])
            cell.alignment = Alignment(horizontal="center")
    
    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Salva arquivo
    wb.save(caminho_excel)

def main():
    caminho_pdf = r"C:\Users\jorge\OneDrive\Ambiente de Trabalho\Cotação - N.00015.pdf"
    caminho_excel = r"C:\Users\jorge\OneDrive\Ambiente de Trabalho\cotacao.xlsx"
    
    texto = extrair_texto_pdf(caminho_pdf)
    produtos = extrair_dados(texto)
    
    if produtos:
        escrever_excel(produtos, caminho_excel)
        print(f"Dados exportados com sucesso para {caminho_excel}")
    else:
        print("Nenhum produto foi encontrado na cotação.")

if __name__ == "__main__":
    main()