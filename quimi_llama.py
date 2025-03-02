import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import os
import sys

class PDFToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor de Cotação PDF para Excel")
        self.root.geometry("440x260")

        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()

        ttk.Label(main_frame, text="Ficheiro PDF de Cotação:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.pdf_path, width=50).grid(row=1, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_pdf).grid(row=1, column=1)

        ttk.Label(main_frame, text="Destino do ficheiro Excel:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.excel_path, width=50).grid(row=3, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_excel).grid(row=3, column=1)

        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.grid(row=4, column=0, columnspan=2, pady=20)

        ttk.Button(main_frame, text="Converter", command=self.convert).grid(row=5, column=0, columnspan=2, pady=10)

        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=6, column=0, columnspan=2)

    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Selecione o ficheiro PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        if filename:
            self.pdf_path.set(filename)
            excel_name = Path(filename).stem + ".xlsx"
            self.excel_path.set(str(Path(filename).with_name(excel_name)))

    def browse_excel(self):
        filename = filedialog.asksaveasfilename(
            title="Salvar ficheiro Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.excel_path.set(filename)

    def extrair_texto_pdf(self, caminho_pdf):
        texto = ""
        with open(caminho_pdf, "rb") as f:
            leitor = PdfReader(f)
            for pagina in leitor.pages:
                texto_extraido = pagina.extract_text()
                if texto_extraido:  # Garantir que não adicionamos None
                    texto += texto_extraido + "\n"
        return texto

    def extrair_dados(self, texto):
        produtos = []

        if "DESCRIÇÃO" in texto:
            texto = texto.split("DESCRIÇÃO", 1)[1]
        else:
            messagebox.showwarning("Aviso", "Cabeçalho 'DESCRIÇÃO' não encontrado.")
            return []

        # Limpar espaços extra
        linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
        texto = "\n".join(linhas)

        padrao = re.compile(
            r"(?P<referencia>\[[^\]]+\])\s*"
            r"(?P<descricao>(?:(?!\[\w+\]).)*?)(?=\s*\d+[.,]\d+(?:\s*(?:[kK][gG]|[lL](?:itros?)?|UN))?)\s*"
            r"(?P<quantidade>\d+[.,]\d+)"
            r"(?:\s*(?P<unidade>(?:[kK][gG]|[lL](?:itros?)?|UN)))?"
            r"\s+(?P<preco>\d+[.,]\d+)"
            r"\s+(?P<impostos>IVA\s*\d+%?)"
            r"\s+(?P<amount>\d+[.,]\d+\s*€)",
            flags=re.DOTALL | re.IGNORECASE
        )

        matches = list(padrao.finditer(texto))

        if not matches:
            messagebox.showwarning("Aviso", "Nenhuma entrada de produto foi encontrada com o padrão definido.")

        for match in matches:
            dados = match.groupdict()

            referencia = dados["referencia"].strip('[]')
            descricao_completa = " ".join(dados["descricao"].split())
            
            descricao_principal = descricao_completa
            descricao_secundaria = ""
            descricao_terciaria = ""
            
            if " Equivalente " in descricao_completa:
                partes_equivalente = descricao_completa.split(" Equivalente ", 1)
                descricao_principal = partes_equivalente[0].strip()
                parte_equivalente = " Equivalente " + partes_equivalente[1].strip()
                
                padrao_divisao = r'(IBC\s.*|Cisterna\s.*|Tambor\s.*|Palete\s.*|Barrica\s.*|Lata\s.*|Jerrican\s.*|TB\s.*)'
                if re.search(padrao_divisao, parte_equivalente, re.IGNORECASE):
                    partes_embalagem = re.split(padrao_divisao, parte_equivalente, 1, flags=re.IGNORECASE)
                    descricao_secundaria = partes_embalagem[0].strip()
                    descricao_terciaria = partes_embalagem[1].strip() if len(partes_embalagem) > 1 else ""
                else:
                    descricao_secundaria = parte_equivalente.strip()
            else:
                padrao_divisao = r'(IBC\s.*|Cisterna\s.*|Tambor\s.*)'
                if re.search(padrao_divisao, descricao_completa, re.IGNORECASE):
                    partes = re.split(padrao_divisao, descricao_completa, 1, flags=re.IGNORECASE)
                    descricao_principal = partes[0].strip()
                    descricao_secundaria = partes[1].strip() if len(partes) > 1 else ""

            quantidade = float(dados["quantidade"].replace(',', '.'))

            unidade = dados.get("unidade", "").upper()
            if unidade.startswith('L'):
                unidade = 'L'
            elif unidade.startswith('K'):
                unidade = 'KG'
            elif unidade == 'UN':
                unidade = 'Unidades'

            preco = float(dados["preco"].replace(',', '.'))

            impostos_text = dados["impostos"]
            num_impostos = re.search(r'\d+[.,]?\d*', impostos_text)
            impostos = float(num_impostos.group().replace(',', '.')) / 100.0 if num_impostos else 0.0

            amount = float(dados["amount"].replace('€', '').replace(',', '.'))

            produtos.append({
                "REFERÊNCIA": referencia,
                "DESCRIÇÃO": descricao_principal,
                "DESCRIÇÃO_SECUNDARIA": descricao_secundaria,
                "DESCRIÇÃO_TERCIARIA": descricao_terciaria,
                "QUANTIDADE": quantidade,
                "UNIDADE": unidade,
                "PREÇO UNITÁRIO": preco,
                "IMPOSTOS": impostos,
                "AMOUNT": amount
            })

        return produtos

    def escrever_excel(self, produtos, caminho_excel):
        workbook = Workbook()
        sheet = workbook.active
        
        headers = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row = 2
        for produto in produtos:
            sheet.cell(row=current_row, column=1).value = produto["REFERÊNCIA"]
            sheet.cell(row=current_row, column=2).value = produto["DESCRIÇÃO"]
            sheet.cell(row=current_row, column=3).value = produto["QUANTIDADE"]
            sheet.cell(row=current_row, column=4).value = produto["UNIDADE"]
            sheet.cell(row=current_row, column=5).value = produto["PREÇO UNITÁRIO"]
            sheet.cell(row=current_row, column=6).value = produto["IMPOSTOS"]
            sheet.cell(row=current_row, column=7).value = produto["AMOUNT"]
            
            if produto["DESCRIÇÃO_SECUNDARIA"]:
                current_row += 1
                sheet.cell(row=current_row, column=2).value = produto["DESCRIÇÃO_SECUNDARIA"]
            
            if produto["DESCRIÇÃO_TERCIARIA"]:
                current_row += 1
                sheet.cell(row=current_row, column=2).value = produto["DESCRIÇÃO_TERCIARIA"]
            
            current_row += 1
        
        for col in range(1, 8):
            sheet.column_dimensions[chr(64 + col)].width = 18
            
        for row in range(2, current_row, 3):
            for col in [3, 5, 6, 7]:
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = "#,##0.00"
        
        workbook.save(caminho_excel)

    def convert(self):
        if not self.pdf_path.get() or not self.excel_path.get():
            messagebox.showerror("Erro", "Por favor, selecione os ficheiros de entrada e saída.")
            return

        try:
            self.status_label.config(text="A processar PDF...")
            self.progress['value'] = 33
            self.root.update_idletasks()

            texto = self.extrair_texto_pdf(self.pdf_path.get())

            self.status_label.config(text="A extrair dados...")
            self.progress['value'] = 66
            self.root.update_idletasks()

            produtos = self.extrair_dados(texto)

            if produtos:
                self.status_label.config(text="A criar ficheiro Excel...")
                self.progress['value'] = 90
                self.root.update_idletasks()

                self.escrever_excel(produtos, self.excel_path.get())

                self.progress['value'] = 100
                self.status_label.config(text="Conversão concluída com sucesso!")
                messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {self.excel_path.get()}")
            else:
                self.status_label.config(text="Nenhum produto encontrado.")
                self.progress['value'] = 0

        except Exception as e:
            self.status_label.config(text="Erro durante a conversão")
            self.progress['value'] = 0
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")

def main():
    root = tk.Tk()
    
    if hasattr(sys, '_MEIPASS'):
        icone_path = Path(sys._MEIPASS) / 'comsoftweb.ico'
    else:
        icone_path = Path(__file__).parent / 'comsoftweb.ico'
    
    try:
        root.iconbitmap(str(icone_path))
    except Exception as e:
        print(f"Aviso: Não foi possível carregar o ícone. Erro: {e}")
    
    app = PDFToExcelConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()
