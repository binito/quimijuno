import re
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def extrair_texto_pdf(caminho_pdf):
    texto = ""
    # Abre o ficheiro PDF em modo leitura binária
    with open(caminho_pdf, "rb") as f:
        leitor = PdfReader(f)
        for pagina in leitor.pages:
            texto += pagina.extract_text() + "\n"
    return texto

def extrair_dados(texto):
    produtos = []
    
    # Separa o texto a partir do cabeçalho "DESCRIÇÃO"
    if "DESCRIÇÃO" in texto:
        texto = texto.split("DESCRIÇÃO", 1)[1]
    else:
        print("Cabeçalho 'DESCRIÇÃO' não encontrado.")
        return []
    
    # Remove linhas vazias e espaços desnecessários
    linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
    texto = "\n".join(linhas)
    
    # Expressão regular atualizada para extrair os campos
    padrao = re.compile(
        r"(?P<referencia>\[[^\]]+\])\s*"  # Captura a referência entre colchetes
        r"(?P<descricao>(?:(?!\[\w+\]).)*?)(?=\s*\d+[.,]\d+(?:\s*(?:[kK][gG]|[lL](?:itros?)?|UN))?)\s*"  # Captura a descrição
        r"(?P<quantidade>\d+[.,]\d+)"     # Captura a quantidade
        r"(?:\s*(?P<unidade>(?:[kK][gG]|[lL](?:itros?)?|UN)))?"  # Captura opcionalmente a unidade
        r"\s+(?P<preco>\d+[.,]\d+)"       # Captura o preço unitário
        r"\s+(?P<impostos>IVA\s*\d+%?)"    # Captura os impostos (ex: IVA 23% ou IVA 23)
        r"\s+(?P<amount>\d+[.,]\d+\s*€)", # Captura o montante total
        flags=re.DOTALL | re.IGNORECASE
    )
    
    matches = list(padrao.finditer(texto))
    
    if not matches:
        print("Nenhuma entrada de produto foi encontrada com o padrão definido.")
    
    for match in matches:
        # Extrai e processa a referência (remove os colchetes)
        referencia = match.group("referencia").strip('[]')
        # Limpa a descrição removendo espaços extra
        descricao = " ".join(match.group("descricao").split())
        
        # Converter a quantidade para float (substituindo a vírgula por ponto)
        quantidade_str = match.group("quantidade").strip()
        quantidade = float(quantidade_str.replace(',', '.'))
        
        # Processa a unidade, se existir
        unidade = match.group("unidade")
        if unidade is None:
            unidade = ""  # Se não detetada, deixa em branco
        else:
            unidade = unidade.upper()
            if unidade.startswith('L'):
                unidade = 'L'
            elif unidade.startswith('K'):
                unidade = 'KG'
            elif unidade == 'UN':
                unidade = 'Unidades'
        
        # Converter o preço unitário para float
        preco_str = match.group("preco").strip()
        preco = float(preco_str.replace(',', '.'))
        
        # Processa os impostos:
        # Extrai o valor numérico (ex: "23" ou "23,5") e converte para float dividindo por 100 
        # para representar a percentagem como decimal (ex: 23 -> 0.23)
        impostos_text = match.group("impostos")
        num_impostos = re.search(r'\d+[.,]?\d*', impostos_text)
        if num_impostos:
            impostos = float(num_impostos.group().replace(',', '.')) / 100.0
        else:
            impostos = 0.0
        
        # Converter o montante total para float (remove o símbolo do euro)
        amount_str = match.group("amount").replace('€','').strip()
        amount = float(amount_str.replace(',', '.'))
        
        produtos.append({
            "REFERÊNCIA": referencia,
            "DESCRIÇÃO": descricao,
            "QUANTIDADE": quantidade,      # Número
            "UNIDADE": unidade,
            "PREÇO UNITÁRIO": preco,         # Número
            "IMPOSTOS": impostos,            # Número (float, ex: 0.23)
            "AMOUNT": amount                 # Número
        })
    
    return produtos

def escrever_excel(produtos, caminho_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotação"
    
    # Define os cabeçalhos das colunas
    cabecalhos = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
    
    # Estilo para os cabeçalhos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Escreve os cabeçalhos na primeira linha
    for col, header in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Escreve os dados dos produtos nas linhas seguintes
    for row, produto in enumerate(produtos, 2):
        for col, header in enumerate(cabecalhos, 1):
            cell = ws.cell(row=row, column=col, value=produto[header])
            cell.alignment = Alignment(horizontal="center")
    
    # Ajusta a largura das colunas com base no comprimento do conteúdo
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Guarda o ficheiro Excel
    wb.save(caminho_excel)

def main():
    # Define os caminhos para o ficheiro PDF e para o ficheiro Excel de saída
    caminho_pdf = r"C:\Users\jorge\OneDrive\Ambiente de Trabalho\Cotação - N.00015.pdf"
    caminho_excel = r"C:\Users\jorge\OneDrive\Ambiente de Trabalho\cotacao.xlsx"
    
    texto = extrair_texto_pdf(caminho_pdf)
    produtos = extrair_dados(texto)
    
    if produtos:
        escrever_excel(produtos, caminho_excel)
        print(f"Dados exportados com sucesso para {caminho_excel}")
    else:
        print("Nenhum produto foi encontrado na cotação.")

if __name__ == "__main__":
    main()
