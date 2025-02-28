import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

class PDFToExcelConverter:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor de Cotação PDF para Excel")
        self.root.geometry("600x300")
        
        # Criar frame principal
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Variáveis para armazenar caminhos
        self.pdf_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        
        # Componentes da GUI
        # PDF Input
        ttk.Label(main_frame, text="Ficheiro PDF de Cotação:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.pdf_path, width=50).grid(row=1, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_pdf).grid(row=1, column=1)
        
        # Excel Output
        ttk.Label(main_frame, text="Destino do ficheiro Excel:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.excel_path, width=50).grid(row=3, column=0, padx=5)
        ttk.Button(main_frame, text="Procurar", command=self.browse_excel).grid(row=3, column=1)
        
        # Progress Bar
        self.progress = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress.grid(row=4, column=0, columnspan=2, pady=20)
        
        # Convert Button
        ttk.Button(main_frame, text="Converter", command=self.convert).grid(row=5, column=0, columnspan=2, pady=10)
        
        # Status Label
        self.status_label = ttk.Label(main_frame, text="")
        self.status_label.grid(row=6, column=0, columnspan=2)

    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Selecione o ficheiro PDF",
            filetypes=[("PDF files", "*.pdf")]
        )
        if filename:
            self.pdf_path.set(filename)
            # Automatically set Excel path
            excel_name = os.path.splitext(os.path.basename(filename))[0] + ".xlsx"
            self.excel_path.set(os.path.join(os.path.dirname(filename), excel_name))

    def browse_excel(self):
        filename = filedialog.asksaveasfilename(
            title="Salvar ficheiro Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.excel_path.set(filename)

    def extrair_texto_pdf(self, caminho_pdf):
        texto = ""
        with open(caminho_pdf, "rb") as f:
            leitor = PdfReader(f)
            for pagina in leitor.pages:
                texto += pagina.extract_text() + "\n"
        return texto

    def extrair_dados(self, texto):
        produtos = []
        
        if "DESCRIÇÃO" in texto:
            texto = texto.split("DESCRIÇÃO", 1)[1]
        else:
            messagebox.showwarning("Aviso", "Cabeçalho 'DESCRIÇÃO' não encontrado.")
            return []
        
        linhas = [linha.strip() for linha in texto.splitlines() if linha.strip()]
        texto = "\n".join(linhas)
        
        padrao = re.compile(
            r"(?P<referencia>\[[^\]]+\])\s*"
            r"(?P<descricao>(?:(?!\[\w+\]).)*?)(?=\s*\d+[.,]\d+\s*(?:[kK][gG]|[lL]|[lL]itros?))\s*"
            r"(?P<quantidade>\d+[.,]\d+)\s*"
            r"(?P<unidade>(?:[kK][gG]|[lL]|[lL]itros?))\s+"
            r"(?P<preco>\d+[.,]\d+)\s+"
            r"(?P<impostos>IVA\s*\d+%?)\s+"
            r"(?P<amount>\d+[.,]\d+\s*€)",
            flags=re.DOTALL | re.IGNORECASE
        )
        
        matches = list(padrao.finditer(texto))
        
        if not matches:
            messagebox.showwarning("Aviso", "Nenhuma entrada de produto foi encontrada com o padrão definido.")
        
        for match in matches:
            referencia = match.group("referencia").strip('[]')
            descricao = " ".join(match.group("descricao").split())
            
            quantidade = match.group("quantidade").strip()
            unidade = match.group("unidade").upper()
            if unidade.startswith('L'):
                unidade = 'L'
            elif unidade.upper().startswith('K'):
                unidade = 'KG'
                
            preco = " ".join(match.group("preco").split())
            
            impostos = match.group("impostos")
            percentagem = re.search(r'\d+%', impostos)
            if percentagem:
                impostos = percentagem.group()
                
            amount = " ".join(match.group("amount").split())
            
            produtos.append({
                "REFERÊNCIA": referencia,
                "DESCRIÇÃO": descricao,
                "QUANTIDADE": quantidade,
                "UNIDADE": unidade,
                "PREÇO UNITÁRIO": preco,
                "IMPOSTOS": impostos,
                "AMOUNT": amount
            })
        
        return produtos

    def escrever_excel(self, produtos, caminho_excel):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cotação"
        
        cabecalhos = ["REFERÊNCIA", "DESCRIÇÃO", "QUANTIDADE", "UNIDADE", "PREÇO UNITÁRIO", "IMPOSTOS", "AMOUNT"]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for row, produto in enumerate(produtos, 2):
            for col, header in enumerate(cabecalhos, 1):
                cell = ws.cell(row=row, column=col, value=produto[header])
                cell.alignment = Alignment(horizontal="center")
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(caminho_excel)

    def convert(self):
        if not self.pdf_path.get() or not self.excel_path.get():
            messagebox.showerror("Erro", "Por favor, selecione os ficheiros de entrada e saída.")
            return
        
        try:
            self.status_label.config(text="A processar PDF...")
            self.progress['value'] = 33
            self.root.update_idletasks()
            
            texto = self.extrair_texto_pdf(self.pdf_path.get())
            
            self.status_label.config(text="A extrair dados...")
            self.progress['value'] = 66
            self.root.update_idletasks()
            
            produtos = self.extrair_dados(texto)
            
            if produtos:
                self.status_label.config(text="A criar ficheiro Excel...")
                self.progress['value'] = 90
                self.root.update_idletasks()
                
                self.escrever_excel(produtos, self.excel_path.get())
                
                self.progress['value'] = 100
                self.status_label.config(text="Conversão concluída com sucesso!")
                messagebox.showinfo("Sucesso", f"Dados exportados com sucesso para {self.excel_path.get()}")
            else:
                self.status_label.config(text="Nenhum produto encontrado.")
                self.progress['value'] = 0
                
        except Exception as e:
            self.status_label.config(text="Erro durante a conversão")
            self.progress['value'] = 0
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")

def main():
    root = tk.Tk()
    app = PDFToExcelConverter(root)
    root.mainloop()

if __name__ == "__main__":
    main()